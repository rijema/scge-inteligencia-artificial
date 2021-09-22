from openpyxl import load_workbook


class InsercaoPergunta:
    
    def insercaoPergunta(self, novaPergunta):
        ## adicionando nova pergunta à base
        workbook = load_workbook(r'.\\Arquivos do Projeto\\BaseReduzida.xlsx')
        nomeDaAba = workbook.sheetnames[0]
        worksheet = workbook[nomeDaAba]
        linhaDaNovaPergunta = worksheet.max_row + 1
        celulaNovaPergunta = worksheet.cell(row=linhaDaNovaPergunta, column=2, value=novaPergunta)
        workbook.save(r'.\\Arquivos do Projeto\\BaseReduzida.xlsx')
        print("Nova pergunta inserida na base para tratamento")
        ##
