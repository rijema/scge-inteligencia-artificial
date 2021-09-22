import pandas as pd
from pandas.core.construction import array
from sklearn.feature_extraction.text import TfidfVectorizer
import numpy as np
from openpyxl import load_workbook
from scipy import sparse

class PreProcessamento:
    
    def preProcessamento(self):

        ## Importando a base da dados já com a nova pergunta inserida
        arquivo = pd.read_excel(r'.\\Arquivos do Projeto\\BaseReduzida.xlsx')

        perguntas = arquivo['PEDIDO']

        ## Aplicando o TF-IDF e vetorizando as perguntas
        vectorizer = TfidfVectorizer()
        X = vectorizer.fit_transform(perguntas)

        arrayDePerguntas = X.toarray()
        perguntaNovaProcessada = arrayDePerguntas[len(arrayDePerguntas) - 1]
        perguntaNovaProcessadaSparse = sparse.csr_matrix(perguntaNovaProcessada)
        perguntaNovaProcessadaSparseArray = perguntaNovaProcessadaSparse.toarray()
        perguntaNovaDf = pd.DataFrame(perguntaNovaProcessadaSparseArray)
        csvPerguntaProcessada = r'.\\Arquivos do Projeto\\perguntaProcessada.csv'
        np.savetxt(csvPerguntaProcessada, perguntaNovaProcessadaSparseArray, delimiter=",")

        ## Salvando as perguntas processadas no formato de vetor 
        #arquivoPreProcessado = r'.\\Arquivos do Projeto\\preProcessadoComPergunta.csv'
        #np.savetxt(arquivoPreProcessado, X.toarray(), delimiter=",")

        ## Excluindo a frase que foi inserida da base
        workbook = load_workbook(r'.\\Arquivos do Projeto\\BaseReduzida.xlsx')
        nomeDaAba = workbook.sheetnames[0]
        worksheet = workbook[nomeDaAba]
        linhaDaNovaPergunta = worksheet.max_row + 1
        worksheet.delete_rows(linhaDaNovaPergunta-1, amount=1)
        workbook.save(r'.\\Arquivos do Projeto\\BaseReduzida.xlsx')

        ## Salvando um arquivo que contém as perguntas no formato de vetor, mas sem a pergunta nova
        arraySemPerguntaNova = arrayDePerguntas[:-1, :]
        arquivoPreProcessado2 = r'.\\Arquivos do Projeto\\preProcessadoSemPergunta.csv'
        np.savetxt(arquivoPreProcessado2, arraySemPerguntaNova, delimiter=",")

        print("Pré processamento executado")
        return perguntaNovaDf